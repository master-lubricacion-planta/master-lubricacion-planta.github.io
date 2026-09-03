# -*- coding: utf-8 -*-
"""Extrae la 'Planilla consumo control de lubricantes V2':
- cons_maestro.json: 35 productos [{id,sku,prod,tipo,uni,visc,min,ubi,envase}]
- cons_stock.json: baseline de stock por id {id:{tam,bin,tin,par,total}} (col TOTAL ACTUAL)
- cons_hist.json: movimientos de la planilla [{f,mov,id,cant,area,eq,resp,obs}]
"""
import openpyxl, json, re, sys, datetime, warnings
warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8')

SRC = r'C:\Users\alexa\Downloads\Planilla consumo control de lubricantes V2 (2).xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)

ws = wb['MAESTRO']
maestro = []
por_sku = {}
por_nombre = {}
for r in range(4, 54):
    prod = ws.cell(row=r, column=3).value
    if not prod or not str(prod).strip(): continue
    it = {
        'id': int(ws.cell(row=r, column=1).value),
        'sku': str(ws.cell(row=r, column=2).value or '').strip(),
        'prod': str(prod).strip(),
        'tipo': str(ws.cell(row=r, column=4).value or '').strip(),
        'uni': str(ws.cell(row=r, column=5).value or '').strip(),
        'visc': str(ws.cell(row=r, column=6).value or '').strip(),
        'min': float(ws.cell(row=r, column=7).value or 0),
        'ubi': str(ws.cell(row=r, column=8).value or '').strip(),
        'envase': str(ws.cell(row=r, column=9).value or '').strip(),
    }
    maestro.append(it)
    if it['sku']: por_sku[it['sku']] = it['id']
    por_nombre[it['prod'].lower()] = it['id']

ws = wb['INVENTARIO']
stock = {}
for r in range(4, 54):
    prod = ws.cell(row=r, column=2).value
    if not prod or not str(prod).strip(): continue
    iid = por_nombre.get(str(prod).strip().lower())
    if iid is None:
        sku = str(ws.cell(row=r, column=3).value or '').strip()
        iid = por_sku.get(sku)
    if iid is None: continue
    g = lambda c: float(ws.cell(row=r, column=c).value or 0)
    stock[str(iid)] = {
        'tam': g(11), 'bin': g(12), 'tin': g(13), 'par': g(14),
        'total': g(15),
    }

def parse_fecha(v):
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date): return v.isoformat()
    s = str(v).strip()
    m = re.fullmatch(r'(\d{4})-(\d{2})-(\d{2}).*', s)
    if m: return f'{m.group(1)}-{m.group(2)}-{m.group(3)}'
    m = re.fullmatch(r'(\d{1,2})-(\d{1,2})-(\d{4})', s)
    if m: return f'{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}'
    return None

ws = wb['CONSUMO']
hist = []
sin_prod = 0
for r in range(4, ws.max_row + 1):
    f = ws.cell(row=r, column=1).value
    if f is None: continue
    fi = parse_fecha(f)
    if not fi: continue
    prod = str(ws.cell(row=r, column=3).value or '').strip()
    sku = str(ws.cell(row=r, column=4).value or '').strip()
    iid = por_nombre.get(prod.lower()) or por_sku.get(sku)
    if iid is None:
        sin_prod += 1
        continue
    # correccion confirmada por Nicolas: las filas '08-11-2026' son del 11 de AGOSTO
    if fi == '2026-11-08': fi = '2026-08-11'
    hist.append({
        'f': fi,
        'mov': str(ws.cell(row=r, column=2).value or '').strip().upper(),
        'id': iid,
        'cant': float(str(ws.cell(row=r, column=6).value or 0).replace(',', '.')),
        'area': str(ws.cell(row=r, column=7).value or '').strip(),
        'eq': str(ws.cell(row=r, column=8).value or '').strip(),
        'resp': str(ws.cell(row=r, column=9).value or '').strip(),
        'obs': str(ws.cell(row=r, column=10).value or '').strip()[:120],
    })

json.dump(maestro, open('cons_maestro.json', 'w', encoding='utf-8'), ensure_ascii=False)
json.dump(stock, open('cons_stock.json', 'w', encoding='utf-8'), ensure_ascii=False)
json.dump(hist, open('cons_hist.json', 'w', encoding='utf-8'), ensure_ascii=False)
import os
print('maestro:', len(maestro), '| stock:', len(stock), '| hist:', len(hist), '| sin match:', sin_prod)
tot_l = sum(s['total'] for k, s in stock.items() if next(m for m in maestro if m['id'] == int(k))['uni'] == 'L')
tot_kg = sum(s['total'] for k, s in stock.items() if next(m for m in maestro if m['id'] == int(k))['uni'] == 'kg')
bajo = sum(1 for k, s in stock.items() if s['total'] < next(m for m in maestro if m['id'] == int(k))['min'])
print(f'stock total: {tot_l:.0f} L aceite | {tot_kg:.0f} kg grasa | bajo minimo: {bajo}')
print('fechas hist:', min(h['f'] for h in hist), '->', max(h['f'] for h in hist))
