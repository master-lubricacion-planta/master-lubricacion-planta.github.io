# -*- coding: utf-8 -*-
"""Extrae una pauta xlsx al formato pautas.json + imagen del equipo.
Uso: python extract_pauta.py <ruta.xlsx> <pln_id>
Agrega/actualiza la entrada en pautas.json y guarda pautas-img/<pln_id>.jpg si hay imagen.
"""
import openpyxl, json, os, sys, io
from PIL import Image

SCRATCH = r'C:\Users\alexa\AppData\Local\Temp\claude\C--Users-alexa-Desktop-Claude-code-Ses-1\ee5231c6-7fdd-4492-a68b-a3a70ea480d3\scratchpad'
PWA = r'C:\Users\alexa\Desktop\Claude code Ses 1\lubricacion-pwa'

def extract(path, pln_id):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    g = lambda r,c: ws.cell(row=r,column=c).value
    title = str(g(2,29) or '').strip()
    equipo = str(g(6,15) or '').strip()
    area = str(g(6,1) or '').strip()
    ubic = str(g(8,1) or '').strip()
    acts = []
    group = ''
    for r in range(21, ws.max_row+1):
        rd = ws.row_dimensions.get(r)
        if rd is not None and rd.hidden:
            continue  # fila oculta: la pauta impresa no la muestra
        c1 = g(r,1); c10 = g(r,10); c26 = g(r,26)
        if c1 and str(c1).startswith('Repuestos'):
            break
        if c1 and not str(c1).strip().isdigit() and not c10:
            group = str(c1).strip()
            continue
        if c1 and str(c1).strip().isdigit() and c10:
            acts.append({'n': int(str(c1).strip()), 'g': group, 'd': str(c10).strip(), 'lim': str(c26 or '').strip()})
    rec = {'id': pln_id, 'titulo': title, 'equipo': equipo, 'area': area, 'ubicacion': ubic, 'acts': acts}

    # imagen: la más grande embebida (fuera del encabezado), aplicando el RECORTE
    # srcRect de Excel — la imagen cruda puede ser una captura de pantalla completa
    # de la que Excel solo muestra una región (la figura del equipo).
    try:
        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.worksheets[0]
        imgs = getattr(ws2, '_images', [])
        best = None; bestpx = 0
        for im in imgs:
            try:
                frm = im.anchor._from
                if frm.row < 5:
                    continue  # logo del encabezado
                data = im._data()
                pil = Image.open(io.BytesIO(data))
                try:
                    src = im.anchor.pic.blipFill.srcRect
                except Exception:
                    src = None
                if src and any([src.l, src.t, src.r, src.b]):
                    W, H = pil.size
                    box = (int(W*(src.l or 0)/100000.0), int(H*(src.t or 0)/100000.0),
                           int(W*(1-(src.r or 0)/100000.0)), int(H*(1-(src.b or 0)/100000.0)))
                    if box[2] > box[0] and box[3] > box[1]:
                        pil = pil.crop(box)
                px = pil.width * pil.height
                if px > bestpx:
                    best, bestpx = pil, px
            except Exception:
                pass
        if best:
            best = best.convert('RGB')
            if best.width > 560:
                best = best.resize((560, int(best.height*560/best.width)), Image.LANCZOS)
            outimg = os.path.join(PWA, 'pautas-img', pln_id + '.jpg')
            best.save(outimg, quality=72, optimize=True)
            has_img = True
        else:
            has_img = False
    except Exception as e:
        has_img = False

    return rec, has_img

if __name__ == '__main__':
    path, pln_id = sys.argv[1], sys.argv[2]
    rec, has_img = extract(path, pln_id)
    pfile = os.path.join(SCRATCH, 'pautas.json')
    pautas = json.load(open(pfile, encoding='utf-8'))
    pautas = [p for p in pautas if p['id'] != pln_id]
    pautas.append(rec)
    json.dump(pautas, open(pfile, 'w', encoding='utf-8'), ensure_ascii=False)
    print(pln_id, '| acts:', len(rec['acts']), '| img:', has_img, '| titulo:', rec['titulo'][:50])
