# -*- coding: utf-8 -*-
"""Extrae el historial de la planilla 'Monitoreo de equipos Críticos 2026' para
la Ruta Crítica: valores por fecha y punto de medición (ids del catálogo RUTA).
Hojas 'Semana 13'..'Semana 35' (layout moderno), secciones ancladas por título.
Salida rprev.json: {fecha_iso: {id: numero | "estado"}}
"""
import openpyxl, json, re, sys, datetime, unicodedata, os
sys.stdout.reconfigure(encoding='utf-8')

SRC = r'C:\Users\alexa\Downloads\Monitoreo de equipos Críticos 2026 (1).xlsx'
ANCHOR = datetime.date(2026, 8, 13)  # jueves semana 33

def fecha(w, di):
    return (ANCHOR + datetime.timedelta(days=(w - 33) * 7 + di)).isoformat()

def nrm(s):
    s = unicodedata.normalize('NFD', str(s or '')).encode('ascii', 'ignore').decode()
    return re.sub(r'\s+', ' ', s).strip().lower()

# (regex_titulo, col_busqueda, tipo, params)
# tipo 'filas': items en offsets desde el titulo, valores en cols c0..c0+6 (un dia por col)
# tipo 'molinos': pares A/B por dia
# tipo 'scrubber': dias como filas t+3..t+9, cols B..F
SEC = [
 (r'^temperatura contraeje', 1, 'filas', dict(c0=2, items={2:'chxce', 3:'chxkg', 4:'chxmc'})),
 (r'registro de presion en bateria', 1, 'filas', dict(c0=2, items={2:'chxpres'})),
 (r'motores correa 1$', 1, 'filas', dict(c0=3, items={2:'c1m1a', 3:'c1m1l', 4:'c1m2a', 5:'c1m2l'})),
 (r'motores correa 2$', 1, 'filas', dict(c0=3, items={2:'c2m1a', 3:'c2m1l', 4:'c2m2a', 5:'c2m2l', 6:'c2m3a', 7:'c2m3l', 8:'c2m4a', 9:'c2m4l'})),
 (r'polea tensora correa 2', 1, 'filas', dict(c0=2, items={2:'c2pd', 3:'c2cd', 4:'c2td', 5:'c2pi', 6:'c2ci', 7:'c2ti'})),
 (r'motores correa 6', 1, 'filas', dict(c0=3, items={2:'c6ma', 3:'c6ml'})),
 (r'motores correa 7', 1, 'filas', dict(c0=3, items={2:'c7ma', 3:'c7ml'})),
 (r'bombas bajo molinos sag', 1, 'filas', dict(c0=3, items={2:'b40b', 3:'b40t', 4:'b41a', 5:'b41l', 6:'b42a', 7:'b42l'})),
 (r'bombas warman', 1, 'filas', dict(c0=3, items={i+2:x for i,x in enumerate(
     ['w11ml','w11ma','w11ac','w11ba','w11bb','w12ml','w12ma','w12ac','w12ba','w12bb',
      'w13ml','w13ma','w13ac','w13ba','w13bb','w14ml','w14ma','w14ac','w14ba','w14bb'])})),
 (r'ventilador scrubber', 1, 'scrubber', dict()),
 (r'presiones filtros de sistema', 11, 'molinos', dict(items={3:('fs1a','fs1b'), 4:('fs2a','fs2b'), 5:('fb1a','fb1b'), 6:('fb2a','fb2b'), 7:('fb3a','fb3b'), 8:('fb4a','fb4b')})),
 (r'aceite en blowers', 11, 'filas', dict(c0=12, items={2:'bl11', 3:'bl12', 4:'bl13'})),
 (r'reductor planta de cal', 11, 'filas', dict(c0=13, items={2:'k243mc', 3:'k243ml', 4:'k243el', 5:'k243re', 6:'k244mc', 7:'k244ml', 8:'k244el', 9:'k244re'})),
 (r'cuerpo de rodad?miento en cal', 11, 'filas', dict(c0=12, items={2:'cal241', 3:'cal242'})),
 (r'espesador 141', 11, 'filas', dict(c0=13, items={2:'e141m1r1', 3:'e141m1r2', 4:'e141m2r1', 5:'e141m2r2', 6:'e141m3r1', 7:'e141m3r2', 8:'e141m4r1', 9:'e141m4r2'})),
 (r'espesador 146', 11, 'filas', dict(c0=13, items={2:'e146m1r1', 3:'e146m1r2', 4:'e146m2r1', 5:'e146m2r2', 6:'e146m3r1', 7:'e146m3r2', 8:'e146m4r1', 9:'e146m4r2'})),
 (r'espesador 281', 11, 'filas', dict(c0=13, items={2:'e281m1r1', 3:'e281m1r2', 4:'e281m2r1', 5:'e281m2r2', 6:'e281m3r1', 7:'e281m3r2', 8:'e281m5r1', 9:'e281m5r2', 10:'e281m6r1', 11:'e281m6r2'})),
 (r'espesador 283', 11, 'filas', dict(c0=13, items={2:'e283m1r1', 3:'e283m1r2', 4:'e283m2r1', 5:'e283m2r2', 6:'e283m3r1', 7:'e283m3r2', 8:'e283m5r1', 9:'e283m5r2', 10:'e283m6r1', 11:'e283m6r2'})),
]

UNIDADES = {'c','oc','°c','ºc','%','kpa','kg','tph','n','amp','hrs','min'}
def parse_val(v):
    """-> ('v', float) | ('e', str) | None"""
    if v is None: return None
    s = str(v).strip()
    if not s: return None
    m = re.search(r'-?\d+(?:[.,]\d+)?', s)
    if m:
        try:
            num = float(m.group(0).replace(',', '.'))
        except ValueError:
            return None
        # descartar celdas que son SOLO la unidad de relleno (' °C', 'kpa'...) sin numero real
        return ('v', num)
    low = nrm(s).replace('º','').replace('°','')
    if low in UNIDADES: return None
    if len(low) < 3: return None
    return ('e', s[:40])

wb = openpyxl.load_workbook(SRC, data_only=True)
prev = {}
usadas = 0
for name in wb.sheetnames:
    m = re.fullmatch(r'Semana\s*0?(\d+)', name.strip())
    if not m: continue
    w = int(m.group(1))
    if w < 13: continue  # layout antiguo, sin extraer
    ws = wb[name]
    # ubicar titulos por columna de busqueda
    titulos = {1: {}, 11: {}}
    for col in (1, 11):
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None: continue
            t = nrm(v)
            if len(t) > 8:
                titulos[col][r] = t
    for pat, col, tipo, prm in SEC:
        rx = re.compile(pat)
        tr = None
        for r, t in titulos[col].items():
            if rx.search(t): tr = r; break
        if tr is None: continue
        if tipo == 'filas':
            for off, iid in prm['items'].items():
                for di in range(7):
                    pv = parse_val(ws.cell(row=tr + off, column=prm['c0'] + di).value)
                    if pv:
                        prev.setdefault(fecha(w, di), {})[iid] = pv[1]
                        usadas += 1
        elif tipo == 'molinos':
            for off, (ida, idb) in prm['items'].items():
                for di in range(7):
                    for k, iid in ((0, ida), (1, idb)):
                        pv = parse_val(ws.cell(row=tr + off, column=12 + 2 * di + k).value)
                        if pv:
                            prev.setdefault(fecha(w, di), {})[iid] = pv[1]
                            usadas += 1
        elif tipo == 'scrubber':
            ids = ['scrllm', 'scrlcm', 'scrltv', 'scrlv', 'scramp']
            for di in range(7):
                for j, iid in enumerate(ids):
                    pv = parse_val(ws.cell(row=tr + 3 + di, column=2 + j).value)
                    if pv:
                        prev.setdefault(fecha(w, di), {})[iid] = pv[1]
                        usadas += 1

json.dump(prev, open('rprev.json', 'w', encoding='utf-8'), ensure_ascii=False)
print('fechas con datos:', len(prev), '| valores:', usadas, '| KB:', os.path.getsize('rprev.json') // 1024)
nums = sum(1 for d in prev.values() for x in d.values() if isinstance(x, (int, float)))
print('numericos:', nums, '| estados:', usadas - nums)
# muestra de validacion contra el dump conocido de S34
f = fecha(34, 1)  # viernes 21-08
d = prev.get(f, {})
print('viernes 21-08:', {k: d[k] for k in ['chxce', 'chxkg', 'c1m1a', 'fs1a', 'e141m1r1'] if k in d})
