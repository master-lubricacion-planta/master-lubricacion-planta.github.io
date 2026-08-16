# -*- coding: utf-8 -*-
"""Genera, por cada pauta xlsx descargada, un HTML fiel al Excel (via xlsx2html)
y un cellmap con las celdas donde inyectar datos del tecnico en runtime.
Cruza contra pautas.json para saber que PLN-IDs ya estan en la biblioteca.
"""
import json, os, re
from collections import defaultdict
import openpyxl
from xlsx2html import xlsx2html
from xlsx2html import core as xlsx2html_core

# Parche: algunas pautas tienen imagenes/formas cuyo anchor.pic es None (bug de xlsx2html
# con ciertos anchors), lo que rompe la conversion completa. Las omitimos en vez de fallar.
_orig_image_to_data = xlsx2html_core.image_to_data
def _safe_images_to_data(ws):
    images_data = defaultdict(list)
    for _i in ws._images:
        try:
            _id = _orig_image_to_data(_i)
        except AttributeError:
            continue
        images_data[(_id["col"], _id["row"])].append(_id)
    return images_data
xlsx2html_core.images_to_data = _safe_images_to_data

SCRATCH = r'C:\Users\alexa\AppData\Local\Temp\claude\C--Users-alexa-Desktop-Claude-code-Ses-1\ee5231c6-7fdd-4492-a68b-a3a70ea480d3\scratchpad'
PWA = r'C:\Users\alexa\Desktop\Claude code Ses 1\lubricacion-pwa'
DL = r'C:\Users\alexa\Downloads'
OUT_HTML_DIR = os.path.join(PWA, 'pautas-html')
os.makedirs(OUT_HTML_DIR, exist_ok=True)

def build_cellmap(ws):
    g = lambda r, c: ws.cell(row=r, column=c).value
    cm = {}

    # xlsx2html solo emite <td id=...> para la celda superior-izquierda de cada rango
    # fusionado; cualquier referencia a una celda "interior" de una fusion no existe en el
    # HTML. Resolvemos siempre a la celda maestra del rango antes de guardarla.
    merged_ranges = list(ws.merged_cells.ranges)
    def master(coord):
        cell = ws[coord]
        for rng in merged_ranges:
            if cell.coordinate in rng:
                return rng.coord.split(':')[0]
        return coord

    # OT: siempre en fila 1, columna donde aparece 'OT' o vacio tras el titulo (AC1 tipicamente)
    ot_cell = None
    for c in range(1, 40):
        v = g(1, c)
        if v and isinstance(v, str) and 'OT' in v.upper():
            ot_cell = ws.cell(row=1, column=c).coordinate
            break
    if not ot_cell:
        ot_cell = 'AC1'
    cm['ot'] = master(ot_cell)

    # Empresa ejecutora / Fecha inicio / Fecha termino: labels en fila 7, valores en fila 8
    label_row = None
    for r in range(6, 9):
        for c in range(1, 40):
            v = g(r, c)
            if v and isinstance(v, str) and 'Empresa Ejecutora' in v:
                label_row = r
                empresa_col = c
                break
        if label_row:
            break
    if label_row:
        cm['empresa'] = master(ws.cell(row=label_row + 1, column=empresa_col).coordinate)
        # fecha inicio: 4 subceldas a la derecha de 'Fecha y Hora Reales - Inicio'
        for c in range(empresa_col + 1, 40):
            v = g(label_row, c)
            if v and isinstance(v, str) and 'Inicio' in v:
                fi_col = c
                cm['fecha_inicio'] = [master(ws.cell(row=label_row + 1, column=fi_col + off).coordinate) for off in (0, 2, 4, 6)]
                break
        for c in range(empresa_col + 1, 40):
            v = g(label_row, c)
            if v and isinstance(v, str) and 'rmino' in v:
                ft_col = c
                cm['fecha_termino'] = [master(ws.cell(row=label_row + 1, column=ft_col + off).coordinate) for off in (0, 2, 4, 6)]
                break

    # tabla de actividades: buscar fila de encabezado con 'N\u00b0' en col A y luego filas con enteros consecutivos
    header_row = None
    for r in range(15, 25):
        if g(r, 1) == 'N\u00b0':
            header_row = r
            break
    valor_col = b_col = m_col = na_col = None
    if header_row:
        for c in range(1, 40):
            v = g(header_row, c)
            if v and isinstance(v, str) and 'Valor' in v:
                valor_col = c
        # fila siguiente suele tener B/M/NA
        for c in range(1, 40):
            v = g(header_row + 1, c)
            if v == 'B':
                b_col = c
            elif v == 'M':
                m_col = c
            elif v == 'NA':
                na_col = c
        if not b_col:
            for c in range(1, 40):
                v = g(header_row, c)
                if v == 'B':
                    b_col = c
                elif v == 'M':
                    m_col = c
                elif v == 'NA':
                    na_col = c

    activities = []
    if header_row and valor_col:
        r = header_row + 2
        max_r = ws.max_row
        while r <= max_r:
            v1 = g(r, 1)
            if isinstance(v1, str) and v1.strip().startswith('Repuestos'):
                break
            if isinstance(v1, (int, float)) and g(r, 10):
                entry = {
                    'n': int(v1),
                    'valor': master(ws.cell(row=r, column=valor_col).coordinate),
                }
                # La columna Estado (B/M/NA) suele venir fusionada en una sola celda por
                # fila (aunque en el encabezado se vean como 3 columnas separadas): basta
                # una celda "estado" donde escribir el texto B/M/NA.
                if b_col:
                    b_master = master(ws.cell(row=r, column=b_col).coordinate)
                    m_master = master(ws.cell(row=r, column=m_col).coordinate) if m_col else None
                    if m_master and m_master == b_master:
                        entry['estado'] = b_master
                    else:
                        entry['b'] = b_master
                        if m_col:
                            entry['m'] = m_master
                        if na_col:
                            entry['na'] = master(ws.cell(row=r, column=na_col).coordinate)
                activities.append(entry)
            r += 1
    cm['activities'] = activities

    # Tecnicos participantes: fila con 'Puesto' y 'Nombre y Apellido'
    tec_header = None
    for r in range(header_row + len(activities) if header_row else 20, min((header_row or 20) + len(activities) + 40, ws.max_row + 1)):
        if g(r, 2) == 'Puesto':
            tec_header = r
            break
        v = g(r, 8)
        if v == 'Nombre y Apellido':
            tec_header = r
            break
    tecnicos = []
    if tec_header:
        # columna de nombre: buscar 'Nombre y Apellido' en esa fila
        nombre_col = None
        for c in range(1, 40):
            v = g(tec_header, c)
            if v == 'Nombre y Apellido':
                nombre_col = c
                break
        if nombre_col:
            for rr in range(tec_header + 1, tec_header + 7):
                tecnicos.append(ws.cell(row=rr, column=nombre_col).coordinate)
    cm['tecnicos'] = tecnicos

    return cm


# Paleta del tema Office estandar, en el orden en que Excel indexa los colores de
# tema dentro del xlsx: 0=lt1(blanco), 1=dk1(negro), 2=lt2, 3=dk2, 4..9=acentos 1-6.
_THEME_RGB = ['FFFFFF', '000000', 'E7E6E6', '44546A', '4472C4', 'ED7D31', 'A5A5A5', 'FFC000', '5B9BD5', '70AD47']

def _tint(rgb_hex, tint):
    """Formula de tinte de Office: negativo oscurece, positivo aclara."""
    out = ''
    for i in (0, 2, 4):
        c = int(rgb_hex[i:i + 2], 16)
        c = c * (1 + tint) if tint < 0 else c * (1 - tint) + 255 * tint
        out += format(max(0, min(255, int(round(c)))), '02X')
    return out

def _fill_rgb(cell):
    """Color de relleno solido de la celda como RRGGBB, o None (los rellenos de tema
    y con tinte los pierde xlsx2html; aqui los resolvemos a mano)."""
    f = cell.fill
    if not f or f.patternType != 'solid':
        return None
    c = f.fgColor
    rgb = None
    try:
        if c.type == 'rgb' and isinstance(c.rgb, str):
            rgb = c.rgb[-6:]
        elif c.type == 'theme' and isinstance(c.theme, int) and c.theme < len(_THEME_RGB):
            rgb = _tint(_THEME_RGB[c.theme], c.tint or 0)
        elif c.type == 'indexed':
            from openpyxl.styles.colors import COLOR_INDEX
            v = COLOR_INDEX[c.indexed]
            rgb = v[-6:]
    except Exception:
        return None
    if not rgb or rgb.upper() in ('FFFFFF', '000000'):
        # blanco = sin relleno visible; 'negro' suele ser el placeholder de "sin color"
        return None
    return rgb


def apply_whitespace(ws, html, sheet_name):
    """Emula el comportamiento real de Excel con el texto:
    - celdas SIN ajustar texto (wrap_text falso): una sola linea que se 'derrama'
      sobre las celdas vecinas vacias (igual que Excel). xlsx2html en cambio las
      envolvia en varias lineas, inflando la altura de las filas (p.ej. la seccion
      SEGURIDAD/EPP quedaba mucho mas alta que en la pauta original).
    - celdas CON ajustar texto: respetar tambien los saltos de linea manuales
      (p.ej. el titulo 'PAUTA DE INSPECCION\nGERENCIA DE MANTENIMIENTO')."""
    html = html.replace('<table  style="border-collapse: collapse"',
                        '<table  style="border-collapse: collapse;table-layout: fixed;'
                        "font-family: Calibri, 'Segoe UI', Arial, sans-serif\"", 1)

    # puede haber otros atributos (rowspan, colspan) entre id y style
    pattern = re.compile(r'(id="' + re.escape(sheet_name) + r'!(?P<coord>[A-Z]+\d+)"[^>]*? style="[^"]*)"')

    def repl(m):
        coord = m.group('coord')
        try:
            cell = ws[coord]
            wrap = cell.alignment.wrap_text
        except Exception:
            cell = None
            wrap = False
        rule = 'white-space:pre-wrap;word-wrap:break-word' if wrap else 'white-space:nowrap'
        if cell is not None:
            rgb = _fill_rgb(cell)
            if rgb:
                rule += ';background-color:#' + rgb
        return m.group(1) + ';' + rule + '"'

    return pattern.sub(repl, html)


def fix_columns(ws, html, sheet_name):
    """Dos defectos de xlsx2html con las columnas:
    1. Las columnas OCULTAS de Excel (p.ej. la Y en las pautas) se omiten del
       <colgroup> pero sus <td> igual se emiten, desplazando el ancho y creando
       una 'columna fantasma' entre Descripcion y Limites. Se ocultan esos td.
    2. Emite <col> extra para columnas con ancho definido mas alla del area usada
       (p.ej. AQ), agrandando la pagina con espacio muerto a la derecha. Se
       recorta el colgroup a las columnas visibles reales."""
    from openpyxl.utils import get_column_letter
    hidden = set()
    for letter, dim in ws.column_dimensions.items():
        if dim.hidden or (dim.width is not None and dim.width == 0):
            hidden.add(letter)

    for letter in hidden:
        pat = re.compile(r'(id="' + re.escape(sheet_name) + '!' + letter + r'\d+"[^>]*? style="[^"]*)"')
        html = pat.sub(lambda m: m.group(1) + ';display:none"', html)

    n_visible = sum(1 for i in range(1, ws.max_column + 1) if get_column_letter(i) not in hidden)
    col_tags = list(re.finditer(r'<col\s+style="width: [\d.]+px">\n?', html))
    if len(col_tags) > n_visible:
        html = html[:col_tags[n_visible].start()] + html[col_tags[-1].end():]
    return html


def inject_figures(ws, html, sheet_name):
    """xlsx2html descarta las imagenes ancladas dentro de rangos fusionados (la celda
    interior no existe como <td> en el HTML). Las inyectamos a mano dentro de la celda
    maestra del rango fusionado donde esta anclada cada imagen (tipicamente FIGURAS)."""
    import base64 as b64
    merged_ranges = list(ws.merged_cells.ranges)

    def master_coord(row, col):
        coord = ws.cell(row=row, column=col).coordinate
        for rng in merged_ranges:
            if coord in rng:
                return rng.coord.split(':')[0]
        return coord

    for im in ws._images:
        try:
            frm = im.anchor._from
        except Exception:
            continue
        if frm.row < 5:
            continue  # el logo del encabezado ya lo maneja xlsx2html
        try:
            data = im._data()
        except Exception:
            continue
        fmt = (im.format or 'png').lower()
        if fmt == 'wmf':
            continue  # openpyxl no soporta wmf, ya viene descartada
        # Excel puede mostrar solo un RECORTE de la imagen embebida (srcRect, en
        # fracciones de 1/100000). Sin aplicarlo se veria la imagen cruda completa
        # (p.ej. una captura de pantalla entera en vez de solo la figura del equipo).
        try:
            src_rect = im.anchor.pic.blipFill.srcRect
        except Exception:
            src_rect = None
        if src_rect and any([src_rect.l, src_rect.t, src_rect.r, src_rect.b]):
            from PIL import Image as PILImage
            import io as _io
            pil = PILImage.open(_io.BytesIO(data))
            W, H = pil.size
            l = (src_rect.l or 0) / 100000.0
            t = (src_rect.t or 0) / 100000.0
            r = (src_rect.r or 0) / 100000.0
            b = (src_rect.b or 0) / 100000.0
            box = (int(W * l), int(H * t), int(W * (1 - r)), int(H * (1 - b)))
            if box[2] > box[0] and box[3] > box[1]:
                pil = pil.crop(box)
                buf = _io.BytesIO()
                pil.convert('RGB').save(buf, 'PNG')
                data = buf.getvalue()
                fmt = 'png'
        target = master_coord(frm.row + 1, frm.col + 1)
        td_id = f'{sheet_name}!{target}'
        marker = f'id="{td_id}"'
        idx = html.find(marker)
        if idx == -1:
            continue
        end_tag = html.find('>', idx)
        if end_tag == -1:
            continue
        # altura de la celda contenedora para escalar la imagen sin desbordar
        # (solo el estilo de ESTE td: desde el id hasta el cierre del tag)
        td_open = html[idx:end_tag]
        mh = re.search(r'height:\s*([\d.]+)pt', td_open)
        h_style = f'max-height:{float(mh.group(1)) - 6:.0f}pt;' if mh else 'max-height:260pt;'
        uri = 'data:image/' + fmt + ';base64,' + b64.b64encode(data).decode()
        img_tag = (f'<img src="{uri}" style="{h_style}max-width:96%;display:block;'
                   f'margin:6pt auto 6pt 30pt;object-fit:contain">')
        html = html[:end_tag + 1] + img_tag + html[end_tag + 1:]
    return html


def main():
    pautas_path = os.path.join(SCRATCH, 'pautas.json')
    pautas = json.load(open(pautas_path, encoding='utf-8'))
    by_id = {p['id']: p for p in pautas}

    processed = 0
    skipped_no_file = []
    errors = []

    ALT_DIR = r'C:\Users\alexa\Desktop\Pautas'
    for pln_id in list(by_id.keys()):
        xlsx_path = os.path.join(DL, pln_id + '.xlsx')
        if not os.path.exists(xlsx_path):
            alt_path = os.path.join(ALT_DIR, pln_id + '.xlsx')
            if os.path.exists(alt_path):
                xlsx_path = alt_path
            else:
                skipped_no_file.append(pln_id)
                continue
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            ws = wb.worksheets[0]
            cellmap = build_cellmap(ws)
            by_id[pln_id]['cellmap'] = cellmap
            by_id[pln_id]['sheet'] = ws.title

            out = xlsx2html(xlsx_path, sheet=0)
            out.seek(0)
            html = out.read()
            html = apply_whitespace(ws, html, ws.title)
            html = fix_columns(ws, html, ws.title)
            html = inject_figures(ws, html, ws.title)
            with open(os.path.join(OUT_HTML_DIR, pln_id + '.html'), 'w', encoding='utf-8') as f:
                f.write(html)
            processed += 1
        except Exception as e:
            errors.append((pln_id, str(e)))

    json.dump(pautas, open(pautas_path, 'w', encoding='utf-8'), ensure_ascii=False)
    print('procesadas:', processed)
    print('sin archivo xlsx local:', len(skipped_no_file), skipped_no_file[:10])
    print('errores:', len(errors))
    for pid, err in errors[:10]:
        print('  ', pid, '->', err[:120])


if __name__ == '__main__':
    main()
