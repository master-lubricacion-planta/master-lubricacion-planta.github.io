# -*- coding: utf-8 -*-
"""Parsea PROGRAMA CON_LUB_Y26W35.xlsx -> plan_w35.json.
Primera pasada de mapeo OT->pauta: match exacto (tag+desc) o (tag solo si el tag
tiene UNA sola pauta en la semana 33 verificada). El resto queda null para
verificar en Maximo."""
import openpyxl, json, re, sys, unicodedata
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')

SRC = r'C:\Users\alexa\Downloads\PROGRAMA CON_LUB_Y26W35.xlsx'
DIAS = ['Jueves','Viernes','Sábado','Domingo','Lunes','Martes','Miércoles']

def norm(s):
    s = str(s or '').replace('\xa0',' ')  # el Excel usa espacios duros
    s = unicodedata.normalize('NFD', s).encode('ascii','ignore').decode()
    return re.sub(r'[^A-Z0-9]+',' ', s.upper()).strip()

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['CON_LUB_Y26W35']

# dias de la semana 35: jueves 27.08 -> miercoles 02.09
import datetime
ini = datetime.date(2026,8,27)
fechas = {(ini+datetime.timedelta(days=i)): DIAS[i] for i in range(7)}
days = [{'nombre':DIAS[i], 'fecha':(ini+datetime.timedelta(days=i)).strftime('%d.%m'), 'ots':[]} for i in range(7)]
idx = {DIAS[i]: i for i in range(7)}

filas = 0; fuera_rango = []
for r in range(6, ws.max_row+1):
    ot = ws.cell(row=r,column=3).value
    if ot is None: continue
    ot = str(ot).strip()
    if not ot.isdigit(): continue
    esp = str(ws.cell(row=r,column=5).value or '').strip()
    desc = str(ws.cell(row=r,column=6).value or '').strip()
    tag = str(ws.cell(row=r,column=7).value or '').strip()
    hini = ws.cell(row=r,column=8).value
    fecha = hini.date() if hasattr(hini,'date') else None
    if fecha not in fechas:
        fuera_rango.append((ot, str(hini))); continue
    days[idx[fechas[fecha]]]['ots'].append({'ot':ot,'tag':tag,'desc':desc,'esp':esp,'pauta':None})
    filas += 1

# --- mapeo primera pasada desde la semana 33 verificada
w33 = json.load(open('plan_w33.json',encoding='utf-8'))
por_tagdesc = {}
por_tag = defaultdict(set)
for d in w33['days']:
    for o in d['ots']:
        if o.get('pauta'):
            por_tagdesc[(norm(o['tag']), norm(o['desc']))] = o['pauta']
            por_tag[norm(o['tag'])].add(o['pauta'])

map_td = map_tag = 0
pend = []
for d in days:
    for o in d['ots']:
        if o['esp'] != 'LUB':
            continue  # no lubricacion: queda sin pauta digital
        k = (norm(o['tag']), norm(o['desc']))
        if k in por_tagdesc:
            o['pauta'] = por_tagdesc[k]; map_td += 1
        else:
            # tag con pauta en s33 pero desc distinta: NO asumir (la pauta depende
            # de la frecuencia/tarea, no solo del equipo) -> verificar en Maximo
            pend.append((d['nombre'], o['ot'], o['tag'], o['desc'][:45].replace('\xa0',' ')))

plan = {'week':35, 'turno':'A', 'anio':2026, 'days':days}
json.dump(plan, open('plan_w35.json','w',encoding='utf-8'), ensure_ascii=False, indent=1)

tot = sum(len(d['ots']) for d in days)
lub = sum(1 for d in days for o in d['ots'] if o['esp']=='LUB')
print(f'OTs: {tot} (LUB: {lub}) | mapeadas tag+desc: {map_td} | tag unico: {map_tag} | PENDIENTES Maximo: {len(pend)}')
print('por dia:', [(d['nombre'], len(d['ots']), sum(1 for o in d['ots'] if o['pauta'])) for d in days])
if fuera_rango: print('fuera de rango:', fuera_rango[:5])
print('\\nPendientes para Maximo:')
for p in pend: print(' ', p)
