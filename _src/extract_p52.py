# -*- coding: utf-8 -*-
"""Extrae del 'Programa de lubricación 52 semanas 2026.xlsx' (hoja Master LUB.):
- p52_items.json: las 340 'Inspección general de lubricación' de frecuencia Semanal
  [{n, area, tag, eq}]
- p52_prev.json: historial ya digitado en la planilla, por semana:
  {"<w>": {"<n>": {"e": "R|C|RE", "d": 0-6|null, "ot": "", "obs": ""}}}
Reglas de lectura por bloque de semana (7 dias + Observaciones + OT):
  celda de dia con R/C/RE -> estado + dia; numero largo -> OT; otro texto -> obs.
"""
import openpyxl, json, re, sys
sys.stdout.reconfigure(encoding='utf-8')

SRC = r'C:\Users\alexa\Downloads\Programa de lubricación 52 semanas 2026.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Master LUB.']

items = []
filas = []
for r in range(8, 1174):
    desc = str(ws.cell(row=r, column=5).value or '').strip()
    freq = str(ws.cell(row=r, column=9).value or '').strip()
    if desc == 'Inspección general de lubricación' and freq == 'Semanal':
        n = ws.cell(row=r, column=1).value
        items.append({
            'n': int(n),
            'area': str(ws.cell(row=r, column=2).value or '').strip(),
            'tag': str(ws.cell(row=r, column=3).value or '').strip(),
            'eq': str(ws.cell(row=r, column=4).value or '').strip(),
        })
        filas.append((int(n), r))

# --- Posiciones REALES de cada bloque semanal, ancladas a los encabezados de la
# fila 6 ('W.k', 'Observaciones', 'OT'). OJO: los bloques NO son uniformes — la
# W.7 mide 11 columnas y la W.20/W.21 miden 10 (columnas insertadas a mano), asi
# que leer a paso fijo de 9 corre todo 4 columnas desde la W.22 (bug real:
# las OT de la W34 aparecian como de la W35).
wpos = {}
for c in range(1, 700):
    v = ws.cell(row=6, column=c).value
    if v is None: continue
    m = re.fullmatch(r'W\.?\s*(\d+)', str(v).strip())
    if m: wpos[int(m.group(1))] = c
assert len(wpos) == 52, f'esperaba 52 encabezados W.k, hay {len(wpos)}'
bloques = {}
for k in range(1, 53):
    ini = wpos[k]
    fin = (wpos[k + 1] - 1) if k < 52 else ini + 8
    dias = list(range(ini, ini + 7))
    obs_c = ot_c = None
    for c in range(ini + 7, fin + 1):
        h = str(ws.cell(row=6, column=c).value or '').strip().lower()
        if h.startswith('observ'): obs_c = c
        elif h == 'ot': ot_c = c
    extras = [c for c in range(ini + 7, fin + 1) if c not in (obs_c, ot_c)]
    bloques[k] = (dias, obs_c, ot_c, extras)

EST = {'R': 'R', 'C': 'C', 'RE': 'RE'}
prev = {}
for n, r in filas:
    for k in range(1, 53):
        dias, obs_c, ot_c, extras = bloques[k]
        e = None; d = None; ot = ''; obs = []
        for i, col in enumerate(dias):
            v = ws.cell(row=r, column=col).value
            if v is None: continue
            s = str(v).strip()
            if not s: continue
            up = s.upper()
            if up in EST and e is None:
                e = EST[up]; d = i
            elif re.fullmatch(r'\d{5,9}', s):
                ot = ot or s
            else:
                obs.append(s)
        for col in ([obs_c, ot_c] + extras):
            if col is None: continue
            v = ws.cell(row=r, column=col).value
            if v is None: continue
            s = str(v).strip()
            if not s: continue
            if re.fullmatch(r'\d{5,9}', s):
                ot = ot or s
            elif s.upper() in EST and e is None:
                e = EST[s.upper()]
            else:
                obs.append(s)
        if e or ot or obs:
            ent = {}
            if e: ent['e'] = e
            if d is not None: ent['d'] = d
            if ot: ent['ot'] = ot
            if obs: ent['obs'] = ' · '.join(dict.fromkeys(obs))[:180]
            prev.setdefault(str(k), {})[str(n)] = ent

json.dump(items, open('p52_items.json', 'w', encoding='utf-8'), ensure_ascii=False)
json.dump(prev, open('p52_prev.json', 'w', encoding='utf-8'), ensure_ascii=False)
import os
tot = sum(len(v) for v in prev.values())
print('items:', len(items), '| semanas con datos:', len(prev), '| entradas historicas:', tot)
print('tamanos: items', os.path.getsize('p52_items.json')//1024, 'KB | prev', os.path.getsize('p52_prev.json')//1024, 'KB')
con_r = sum(1 for w in prev.values() for x in w.values() if x.get('e')=='R')
print('con R:', con_r)
