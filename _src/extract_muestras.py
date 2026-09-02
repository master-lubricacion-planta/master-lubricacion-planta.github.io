# -*- coding: utf-8 -*-
"""Extrae del 'Programa Toma de Muestras Concentradora 2026' (hoja Calendario 2026):
- mu_items.json: 317 puntos de muestreo [{n, area, tag, eq, lub}]
- mu_prev.json: historial por mes {"1".."12": {"<n>": {m:'R'|'NT', d:0-6, rc, rep, est}}}
Bloques mensuales anclados a los rotulos de la fila 5 (leccion del Plan 52).
Cada bloque: 7 dias (J V S D L M MI) + 'Relleno o cambio' + 'Reporte de analisis' + 'Estado'.
"""
import openpyxl, json, re, sys
sys.stdout.reconfigure(encoding='utf-8')

SRC = r'C:\Users\alexa\Downloads\Programa Toma de Muestras Concentradora 2026 (1).xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Calendario 2026']
MES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

pos = {}
for c in range(1, 140):
    v = str(ws.cell(row=5, column=c).value or '').strip()
    if v in MES: pos[MES.index(v) + 1] = c
assert len(pos) == 12, pos

items = []
filas = []
for r in range(7, 324):
    tag = str(ws.cell(row=r, column=3).value or '').strip()
    if not tag: continue
    n = int(ws.cell(row=r, column=1).value)
    items.append({
        'n': n,
        'area': str(ws.cell(row=r, column=2).value or '').strip(),
        'tag': tag,
        'eq': str(ws.cell(row=r, column=5).value or '').strip(),
        'lub': str(ws.cell(row=r, column=7).value or '').strip(),
    })
    filas.append((n, r))

prev = {}
for n, r in filas:
    for mes, b in pos.items():
        m = None; d = None; rc = ''; rep = ''; est = ''; extra = []
        for i in range(7):
            v = ws.cell(row=r, column=b + i).value
            if v is None: continue
            s = str(v).strip()
            if not s: continue
            up = s.upper()
            if up in ('R', 'NT') and m is None:
                m = up; d = i
            elif re.fullmatch(r'\d{6,12}', s):
                rep = rep or s
            else:
                extra.append(s)
        v = ws.cell(row=r, column=b + 7).value   # Relleno o cambio
        if v is not None and str(v).strip(): rc = str(v).strip()[:20]
        v = ws.cell(row=r, column=b + 8).value   # Reporte de analisis
        if v is not None and str(v).strip():
            s = str(v).strip()
            if re.fullmatch(r'\d[\d-]{3,15}', s):
                rep = rep or s
            else:
                extra.append(s)
        v = ws.cell(row=r, column=b + 9).value   # Estado del analisis
        if v is not None and str(v).strip(): est = str(v).strip()[:15]
        if m or rc or rep or est or extra:
            ent = {}
            if m: ent['m'] = m
            if d is not None: ent['d'] = d
            if rc: ent['rc'] = rc
            if rep: ent['rep'] = rep
            if est: ent['est'] = est
            if extra: ent['obs'] = ' · '.join(dict.fromkeys(extra))[:150]
            prev.setdefault(str(mes), {})[str(n)] = ent

json.dump(items, open('mu_items.json', 'w', encoding='utf-8'), ensure_ascii=False)
json.dump(prev, open('mu_prev.json', 'w', encoding='utf-8'), ensure_ascii=False)
import os
print('items:', len(items), '| meses con datos:', sorted(int(k) for k in prev))
for k in sorted(prev, key=int):
    v = prev[k]
    print(' mes', k, '-> R:', sum(1 for x in v.values() if x.get('m') == 'R'),
          '| NT:', sum(1 for x in v.values() if x.get('m') == 'NT'), '| total:', len(v))
print('tamanos:', os.path.getsize('mu_items.json') // 1024, 'KB /', os.path.getsize('mu_prev.json') // 1024, 'KB')
